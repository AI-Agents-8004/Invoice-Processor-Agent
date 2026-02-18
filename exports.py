"""
Export helpers — convert InvoiceData into CSV or Excel (bytes).
"""

import csv
import io
from models import InvoiceData


# ── CSV ────────────────────────────────────────────────────────────────────────

def to_csv(data: InvoiceData, invoice_id: str) -> bytes:
    """
    Returns a UTF-8 CSV as bytes.

    Layout:
      Section 1 — Invoice summary (label, value rows)
      Section 2 — Line items table
    """
    buf = io.StringIO()
    writer = csv.writer(buf)

    # ── Header banner ──────────────────────────────────────────────────────────
    writer.writerow(["INVOICE DATA EXPORT"])
    writer.writerow(["Invoice ID", invoice_id])
    writer.writerow([])

    # ── Summary section ────────────────────────────────────────────────────────
    writer.writerow(["=== VENDOR ==="])
    writer.writerow(["Vendor Name",          data.vendor_name or ""])
    writer.writerow(["Vendor Address",        data.vendor_address or ""])
    writer.writerow(["Vendor Email",          data.vendor_email or ""])
    writer.writerow(["Vendor Phone",          data.vendor_phone or ""])
    writer.writerow(["Vendor Tax ID",         data.vendor_tax_id or ""])
    writer.writerow([])

    writer.writerow(["=== CLIENT ==="])
    writer.writerow(["Client Name",           data.client_name or ""])
    writer.writerow(["Client Address",        data.client_address or ""])
    writer.writerow(["Client Email",          data.client_email or ""])
    writer.writerow([])

    writer.writerow(["=== INVOICE META ==="])
    writer.writerow(["Invoice Number",        data.invoice_number or ""])
    writer.writerow(["Invoice Date",          data.invoice_date or ""])
    writer.writerow(["Due Date",              data.due_date or ""])
    writer.writerow(["PO Number",             data.purchase_order_number or ""])
    writer.writerow(["Currency",              data.currency or ""])
    writer.writerow([])

    writer.writerow(["=== TOTALS ==="])
    writer.writerow(["Subtotal",              data.subtotal or ""])
    writer.writerow(["Tax Rate (%)",          data.tax_rate or ""])
    writer.writerow(["Tax Amount",            data.tax_amount or ""])
    writer.writerow(["Discount",              data.discount or ""])
    writer.writerow(["Shipping",              data.shipping or ""])
    writer.writerow(["TOTAL AMOUNT",          data.total_amount or ""])
    writer.writerow([])

    writer.writerow(["=== PAYMENT ==="])
    writer.writerow(["Payment Terms",         data.payment_terms or ""])
    writer.writerow(["Payment Method",        data.payment_method or ""])
    writer.writerow(["Bank Account",          data.bank_account or ""])
    writer.writerow([])

    if data.notes:
        writer.writerow(["=== NOTES ==="])
        writer.writerow(["Notes", data.notes])
        writer.writerow([])

    # ── Line items section ─────────────────────────────────────────────────────
    writer.writerow(["=== LINE ITEMS ==="])
    writer.writerow(["#", "Description", "Quantity", "Unit Price", "Total"])
    for i, item in enumerate(data.line_items, start=1):
        writer.writerow([
            i,
            item.description,
            item.quantity if item.quantity is not None else "",
            item.unit_price if item.unit_price is not None else "",
            item.total if item.total is not None else "",
        ])

    return buf.getvalue().encode("utf-8-sig")   # utf-8-sig = BOM for Excel compatibility


# ── Excel ──────────────────────────────────────────────────────────────────────

def to_excel(data: InvoiceData, invoice_id: str) -> bytes:
    """
    Returns an .xlsx file as bytes with two sheets:
      Sheet 1 — Invoice Summary
      Sheet 2 — Line Items
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: Summary ───────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Invoice Summary"

    BLUE   = "2563EB"
    LBLUE  = "DBEAFE"
    DGRAY  = "1E293B"
    LGRAY  = "F8FAFC"

    def header_style(cell, text: str):
        cell.value = text
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    def label_style(cell, text: str):
        cell.value = text
        cell.font = Font(bold=True, color=DGRAY, size=10)
        cell.fill = PatternFill("solid", fgColor=LBLUE)

    def value_style(cell, value):
        cell.value = value
        cell.font = Font(color=DGRAY, size=10)
        cell.fill = PatternFill("solid", fgColor=LGRAY)

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def add_section(ws, title: str, rows: list[tuple], start_row: int) -> int:
        """Add a titled section. Returns next available row."""
        header_cell = ws.cell(row=start_row, column=1, value=title)
        header_style(header_cell, title)
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
        header_cell.border = border
        ws.cell(row=start_row, column=2).border = border

        for label, value in rows:
            start_row += 1
            lc = ws.cell(row=start_row, column=1)
            vc = ws.cell(row=start_row, column=2)
            label_style(lc, label)
            value_style(vc, value)
            lc.border = border
            vc.border = border

        return start_row + 2   # blank gap after section

    # Title row
    title_cell = ws1.cell(row=1, column=1, value="INVOICE EXPORT")
    title_cell.font = Font(bold=True, size=14, color=BLUE)
    ws1.cell(row=1, column=2, value=f"ID: {invoice_id}").font = Font(size=10, color="64748B")
    ws1.row_dimensions[1].height = 24

    row = 3

    row = add_section(ws1, "VENDOR", [
        ("Vendor Name",    data.vendor_name),
        ("Address",        data.vendor_address),
        ("Email",          data.vendor_email),
        ("Phone",          data.vendor_phone),
        ("Tax ID",         data.vendor_tax_id),
    ], row)

    row = add_section(ws1, "CLIENT", [
        ("Client Name",    data.client_name),
        ("Address",        data.client_address),
        ("Email",          data.client_email),
    ], row)

    row = add_section(ws1, "INVOICE META", [
        ("Invoice Number", data.invoice_number),
        ("Invoice Date",   data.invoice_date),
        ("Due Date",       data.due_date),
        ("PO Number",      data.purchase_order_number),
        ("Currency",       data.currency),
    ], row)

    row = add_section(ws1, "TOTALS", [
        ("Subtotal",       data.subtotal),
        ("Tax Rate (%)",   data.tax_rate),
        ("Tax Amount",     data.tax_amount),
        ("Discount",       data.discount),
        ("Shipping",       data.shipping),
        ("TOTAL AMOUNT",   data.total_amount),
    ], row)

    row = add_section(ws1, "PAYMENT", [
        ("Payment Terms",  data.payment_terms),
        ("Payment Method", data.payment_method),
        ("Bank Account",   data.bank_account),
    ], row)

    if data.notes:
        add_section(ws1, "NOTES", [("Notes", data.notes)], row)

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 45

    # ── Sheet 2: Line Items ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Line Items")

    col_headers = ["#", "Description", "Quantity", "Unit Price", "Total"]
    col_widths   = [5,   55,            12,         14,           14]

    for col_idx, (h, w) in enumerate(zip(col_headers, col_widths), start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws2.column_dimensions[get_column_letter(col_idx)].width = w

    ws2.row_dimensions[1].height = 20

    for row_idx, item in enumerate(data.line_items, start=2):
        fill = PatternFill("solid", fgColor="F1F5F9" if row_idx % 2 == 0 else "FFFFFF")
        values = [row_idx - 1, item.description, item.quantity, item.unit_price, item.total]
        for col_idx, val in enumerate(values, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(size=10)
            cell.fill = fill
            cell.border = border
            if col_idx in (3, 4, 5):
                cell.alignment = Alignment(horizontal="right")
                if val is not None:
                    cell.number_format = "#,##0.00"

    # Totals row at bottom
    if data.line_items:
        total_row = len(data.line_items) + 2
        ws2.cell(row=total_row, column=1, value="")
        ws2.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
        ws2.cell(row=total_row, column=5, value=data.total_amount).font = Font(bold=True)
        ws2.cell(row=total_row, column=5).number_format = "#,##0.00"

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
